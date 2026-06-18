"""
Tests HTTP del soporte Excel de Bit2Me (detrás del flag BIT2ME_EXCEL_ENABLED).

Cubre:
  · flag OFF → comportamiento actual intacto (solo CSV; endpoint años 404)
  · flag ON  → /api/bit2me/anos y /api/analizar procesan el Excel vía MotorFIFO
  · permuta con carryover de coste
  · smoke con el Excel real (si está disponible)
"""

import io
import os
import sys
import tempfile
from datetime import datetime

import openpyxl
import pytest

_TEST_DB = tempfile.NamedTemporaryFile(suffix=".db", delete=False).name
os.environ.setdefault("DATABASE_URL", f"sqlite:///{_TEST_DB}")
os.environ.setdefault("SECRET_KEY", "test-secret-bit2me-excel-http")

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import app as appmod
from app import app as flask_app, db
from models import User

HEADERS = [
    "Tipo de operación", "Cantidad de destino", "Moneda de destino",
    "Cantidad de origen", "Moneda de origen", "Comisión de la operación",
    "Moneda de la comisión", "Exchange", "Grupo", "Descripción", "Fecha",
]

def _row(tipo, cd, md, co, mo, fee=0, fecha="2025-03-01 10:00"):
    return [tipo, cd, md, co, mo, fee, "EUR", "Bit2Me", "pocket", "desc", fecha]

# Dataset representativo (2025): compra, venta, permuta, staking, deposit cripto, withdrawal
_ROWS = [
    _row("Trade", 87.07, "XRP", 196.02, "EUR", fee=1.86, fecha="2025-01-07 19:26"),
    _row("Trade", 30.0,  "EUR", 50.0,  "XRP", fee=0.1,  fecha="2025-06-20 13:29"),
    _row("Trade", 22.79, "XRP", 148.40, "ALGO", fee=0.48, fecha="2025-01-08 16:43"),
    _row("Staking", 1.99, "B2M", None, None, fecha="2025-01-01 11:02"),
    _row("Deposit", 10, "XLM", 10, "XLM", fecha="2025-01-09 14:05"),
    _row("Withdrawal", 11.5, "XRP", 11.5, "XRP", fecha="2025-02-05 22:35"),
]


def _xlsx(rows) -> io.BytesIO:
    wb = openpyxl.Workbook(); ws = wb.active
    ws.cell(row=1, column=1, value=" ")
    for j, h in enumerate(HEADERS, start=1):
        ws.cell(row=2, column=j, value=h)
    for i, fila in enumerate(rows, start=3):
        for j, v in enumerate(fila, start=1):
            if v is not None:
                ws.cell(row=i, column=j, value=v)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# CSV "Informe Fiscal" mínimo (formato pre-FIFO con "Yes") — para verificar no regresión.
_CSV_INFORME = "".join(f'"{v}"\n' for v in [
    "Bit2Me Informe Fiscal", "Estimado usuario", "12345",
    "10.01.2024 09:00:00", "Buy", "0,5", "BTC", "20000", "EUR",
    "18000", "20000", "2000", "Yes",
])


@pytest.fixture(scope="module")
def _app():
    flask_app.config.update(TESTING=True, RATELIMIT_ENABLED=False,
                            WTF_CSRF_ENABLED=False, SESSION_COOKIE_SECURE=False,
                            SQLALCHEMY_ENGINE_OPTIONS={})
    with flask_app.app_context():
        db.create_all()
        yield flask_app
        db.session.remove(); db.drop_all()


@pytest.fixture
def client(_app):
    return _app.test_client()


@pytest.fixture
def usuario(_app):
    with _app.app_context():
        u = User(email="b2mx@e.com", full_name="B", email_verified_at=datetime.utcnow())
        u.set_password("p"); db.session.add(u); db.session.commit(); uid = u.id
    yield uid
    with _app.app_context():
        User.query.filter_by(id=uid).delete(); db.session.commit()


@pytest.fixture
def flag_on(monkeypatch):
    monkeypatch.setattr(appmod, "_BIT2ME_EXCEL_ENABLED", True)


@pytest.fixture
def flag_off(monkeypatch):
    monkeypatch.setattr(appmod, "_BIT2ME_EXCEL_ENABLED", False)


def _login(client):
    client.post("/api/login", json={"email": "b2mx@e.com", "password": "p"})
    return client


# ── FLAG OFF: comportamiento actual intacto ───────────────────────────────────

def test_off_endpoint_anos_devuelve_404(client, usuario, flag_off):
    _login(client)
    r = client.post("/api/bit2me/anos",
                    data={"file": (_xlsx(_ROWS), "bit2me.xlsx")},
                    content_type="multipart/form-data")
    assert r.status_code == 404

def test_off_analizar_rechaza_xlsx_bit2me(client, usuario, flag_off):
    _login(client)
    r = client.post("/api/analizar",
                    data={"csv": (_xlsx(_ROWS), "bit2me.xlsx"), "exchange": "bit2me",
                          "ejercicio": "2025"},
                    content_type="multipart/form-data")
    assert r.status_code == 400
    assert ".csv" in r.get_json()["error"]

def test_off_csv_informe_fiscal_sigue_funcionando(client, usuario, flag_off):
    _login(client)
    r = client.post("/api/analizar",
                    data={"csv": (io.BytesIO(_CSV_INFORME.encode()), "informe.csv"),
                          "exchange": "bit2me", "ejercicio": "2024", "nombre": "B"},
                    content_type="multipart/form-data")
    assert r.status_code == 200
    body = r.get_json()
    assert body["ok"] is True
    assert body["resumen"]["operaciones_con_resultado"] == 1
    assert body.get("estimacion") is False


# ── FLAG ON: el Excel funciona ────────────────────────────────────────────────

def test_on_anos_detecta_2025(client, usuario, flag_on):
    _login(client)
    r = client.post("/api/bit2me/anos",
                    data={"file": (_xlsx(_ROWS), "bit2me.xlsx")},
                    content_type="multipart/form-data")
    body = r.get_json()
    assert body["ok"] is True and body["anos"] == [2025]

def test_on_analizar_acepta_xlsx_y_devuelve_resumen_motor(client, usuario, flag_on):
    _login(client)
    r = client.post("/api/analizar",
                    data={"csv": (_xlsx(_ROWS), "bit2me.xlsx"), "exchange": "bit2me",
                          "ejercicio": "2025", "nombre": "B"},
                    content_type="multipart/form-data")
    assert r.status_code == 200
    body = r.get_json()
    assert body["ok"] is True
    assert body["estimacion"] is True
    assert "token" in body and body["token"]
    assert any("ORIENTATIVO" in a for a in body["advertencias"])
    assert "operaciones_con_resultado" in body["resumen"]

def test_on_permuta_no_genera_ganancia_por_carryover(client, usuario, flag_on):
    _login(client)
    # Solo una permuta: el swap cristaliza ganancia 0 (coste se traspasa)
    rows = [_row("Trade", 100, "XRP", 100, "ALGO", fecha="2025-03-01 10:00")]
    r = client.post("/api/analizar",
                    data={"csv": (_xlsx(rows), "bit2me.xlsx"), "exchange": "bit2me",
                          "ejercicio": "2025", "nombre": "B"},
                    content_type="multipart/form-data")
    body = r.get_json()
    assert body["ok"] is True
    assert body["resumen"]["ganancias_brutas"] == 0
    assert body["resumen"]["perdidas_brutas"] == 0

def test_on_rechaza_extension_no_soportada(client, usuario, flag_on):
    _login(client)
    r = client.post("/api/analizar",
                    data={"csv": (io.BytesIO(b"x"), "f.txt"), "exchange": "bit2me",
                          "ejercicio": "2025"},
                    content_type="multipart/form-data")
    assert r.status_code == 400


# ── SMOKE con el Excel real (si está disponible) ──────────────────────────────

_REAL = os.path.expanduser("~/Downloads/bit2me.xlsx")

@pytest.mark.skipif(not os.path.exists(_REAL), reason="Excel real no disponible")
def test_smoke_excel_real(client, usuario, flag_on):
    _login(client)
    # detección de años
    with open(_REAL, "rb") as fh:
        r = client.post("/api/bit2me/anos", data={"file": (fh, "bit2me.xlsx")},
                        content_type="multipart/form-data")
    assert r.get_json()["ok"] is True
    assert 2025 in r.get_json()["anos"]
    # análisis + PDF
    with open(_REAL, "rb") as fh:
        r = client.post("/api/analizar",
                        data={"csv": (fh, "bit2me.xlsx"), "exchange": "bit2me",
                              "ejercicio": "2025", "nombre": "Real"},
                        content_type="multipart/form-data")
    body = r.get_json()
    assert body["ok"] is True and body["estimacion"] is True
    assert body["token"]
    assert len(body["operaciones"]) > 0
