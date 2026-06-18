"""
Clasificador del Excel "Historial de operaciones" de Bit2Me.

Verifica el mapeo columna→dataclass, el parsing defensivo y las advertencias,
sobre un fixture que replica el esquema real (anonimizado).
"""

import os
import tempfile
from datetime import datetime

import openpyxl
import pytest

from clasificador_bit2me_excel import (
    ClasificadorBit2MeExcel,
    validar_columnas_bit2me_excel,
    Bit2MeExcelError,
    _parse_decimal,
    _parse_fecha,
)

HEADERS = [
    "Tipo de operación", "Cantidad de destino", "Moneda de destino",
    "Cantidad de origen", "Moneda de origen", "Comisión de la operación",
    "Moneda de la comisión", "Exchange", "Grupo", "Descripción", "Fecha",
]


def _build_xlsx(rows, *, header_junk=True) -> str:
    """Crea un xlsx temporal con cabecera + filas. fd se cierra al volver."""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb = openpyxl.Workbook()
    ws = wb.active
    r = 1
    if header_junk:
        ws.cell(row=r, column=1, value=" "); r += 1   # 1ª fila basura como el real
    for j, h in enumerate(HEADERS, start=1):
        ws.cell(row=r, column=j, value=h)
    r += 1
    for fila in rows:
        for j, v in enumerate(fila, start=1):
            if v is not None:
                ws.cell(row=r, column=j, value=v)
        r += 1
    wb.save(path)
    return path


def _row(tipo, cd=None, md=None, co=None, mo=None, fee=0, grupo="", desc="",
         fecha="2025-03-01 10:00"):
    return [tipo, cd, md, co, mo, fee, "EUR", "Bit2Me", grupo, desc, fecha]


# ── Helpers de parsing ────────────────────────────────────────────────────────

def test_parse_fecha_sin_segundos_se_normaliza():
    assert _parse_fecha("2025-01-07 19:26") == "2025-01-07 19:26:00"

def test_parse_fecha_datetime_nativo():
    assert _parse_fecha(datetime(2025, 1, 7, 19, 26)) == "2025-01-07 19:26:00"

def test_parse_fecha_invalida_devuelve_none():
    assert _parse_fecha("no-es-fecha") is None
    assert _parse_fecha(None) is None

def test_parse_decimal_nativo_y_europeo():
    assert float(_parse_decimal(196.02)) == 196.02
    assert float(_parse_decimal("1.234,56")) == 1234.56
    assert float(_parse_decimal("196.02")) == 196.02

def test_parse_decimal_fee_malformado_es_none():
    assert _parse_decimal("0.99500.46554750") is None


# ── Mapeo por tipo ────────────────────────────────────────────────────────────

def _clasificar(rows):
    p = _build_xlsx(rows)
    try:
        return ClasificadorBit2MeExcel(p).clasificar()
    finally:
        os.unlink(p)

def test_trade_eur_a_cripto_es_compra():
    c = _clasificar([_row("Trade", 87.07, "XRP", 196.02, "EUR", fee=1.86)])
    assert len(c.compraventas) == 1 and not c.swaps
    op = c.compraventas[0]
    assert op.tipo == "COMPRA" and op.activo == "XRP" and op.contraparte == "EUR"
    assert op.cantidad == 87.07 and op.importe == 196.02 and op.fee_cantidad == 1.86

def test_trade_cripto_a_eur_es_venta():
    c = _clasificar([_row("Trade", 7.70, "EUR", 6358.4, "SBR", fee=0.07)])
    op = c.compraventas[0]
    assert op.tipo == "VENTA" and op.activo == "SBR"
    assert op.cantidad == 6358.4 and op.importe == 7.70

def test_trade_cripto_a_cripto_es_swap():
    c = _clasificar([_row("Trade", 22.79, "XRP", 148.40, "ALGO", fee=0.48)])
    assert not c.compraventas and len(c.swaps) == 1
    s = c.swaps[0]
    assert s.activo_entregado == "ALGO" and s.cantidad_entregada == 148.40
    assert s.activo_recibido == "XRP" and s.cantidad_recibida == 22.79

def test_staking_es_rendimiento_sin_valor_eur():
    c = _clasificar([_row("Staking", 1.99, "B2M", None, None, fecha="2025-01-01 11:02")])
    assert len(c.rendimientos) == 1
    r = c.rendimientos[0]
    assert r.subtipo == "Staking" and r.activo == "B2M" and r.cantidad == 1.99
    assert not hasattr(r, "valor_eur")

def test_other_income_es_rendimiento():
    c = _clasificar([_row("Other Income", 2, "EURC", 2, "EURC")])
    assert len(c.rendimientos) == 1 and c.rendimientos[0].subtipo == "Other Income"

def test_deposit_eur_es_movimiento_sin_inventario():
    c = _clasificar([_row("Deposit", 196.02, "EUR", 200, "EUR", fee=3.98, grupo="card")])
    assert not c.compraventas
    assert len(c.movimientos) == 1 and c.movimientos[0].subtipo == "Deposit EUR"

def test_deposit_cripto_es_compra_importe_cero_con_advertencia():
    c = _clasificar([_row("Deposit", 10, "XLM", 10, "XLM", grupo="blockchain")])
    assert len(c.compraventas) == 1
    op = c.compraventas[0]
    assert op.tipo == "COMPRA" and op.activo == "XLM" and op.importe == 0.0
    assert any("sin coste de adquisición" in a for a in c.advertencias)

def test_withdrawal_cripto_es_movimiento():
    c = _clasificar([_row("Withdrawal", 11.5, "XRP", 11.5, "XRP", grupo="blockchain")])
    assert len(c.movimientos) == 1 and c.movimientos[0].subtipo == "Withdrawal"
    assert any("retiradas de cripto" in a for a in c.advertencias)

def test_withdrawal_eur_es_movimiento():
    c = _clasificar([_row("Withdrawal", 50, "EUR", 50, "EUR")])
    assert len(c.movimientos) == 1 and c.movimientos[0].subtipo == "Withdrawal EUR"

def test_fee_malformado_se_trata_como_cero_con_advertencia():
    c = _clasificar([_row("Trade", 17.0, "XRP", 50, "EUR", fee="0.99500.46554750")])
    assert c.compraventas[0].fee_cantidad == 0.0
    assert any("comisión" in a.lower() for a in c.advertencias)

def test_tipo_desconocido_va_a_desconocidas():
    c = _clasificar([_row("Frobnicate", 1, "FOO", 1, "FOO")])
    assert len(c.desconocidas) == 1 and "tipo:" in c.desconocidas[0].subtipo

def test_fila_fecha_vacia_va_a_desconocidas():
    c = _clasificar([_row("Trade", 1, "XRP", 1, "EUR", fecha=None)])
    assert len(c.desconocidas) == 1 and c.desconocidas[0].subtipo == "fecha_invalida"

def test_advertencia_historial_previo_siempre():
    c = _clasificar([_row("Trade", 1, "XRP", 10, "EUR")])
    assert any("acumulativo" in a for a in c.advertencias)

def test_interfaz_compatible_con_pipeline():
    c = _clasificar([_row("Trade", 1, "XRP", 10, "EUR")])
    for attr in ("compraventas", "swaps", "rendimientos", "movimientos",
                 "desconocidas", "advertencias"):
        assert hasattr(c, attr)


# ── Validación de columnas ────────────────────────────────────────────────────

def test_validar_columnas_ok():
    p = _build_xlsx([_row("Trade", 1, "XRP", 10, "EUR")])
    try:
        ok, _ = validar_columnas_bit2me_excel(p)
        assert ok
    finally:
        os.unlink(p)

def test_validar_columnas_rechaza_fichero_sin_cabecera():
    fd, p = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
    wb = openpyxl.Workbook(); ws = wb.active
    ws.cell(row=1, column=1, value="cualquier cosa"); wb.save(p)
    try:
        ok, msg = validar_columnas_bit2me_excel(p)
        assert not ok and "Bit2Me" in msg
    finally:
        os.unlink(p)

def test_clasificar_fichero_invalido_lanza_error_tipado():
    fd, p = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
    wb = openpyxl.Workbook(); ws = wb.active
    ws.cell(row=1, column=1, value="nada"); wb.save(p)
    try:
        with pytest.raises(Bit2MeExcelError) as exc:
            ClasificadorBit2MeExcel(p).clasificar()
        assert exc.value.category == "user_error"
    finally:
        os.unlink(p)
