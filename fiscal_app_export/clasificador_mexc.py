"""
Clasificador fiscal de operaciones MEXC
Mariano Sevilla — marianosevilla.com

Soporta los exports XLS/XLSX de MEXC:
  · Spot Trade Records  (columnas: order_id, symbol, currency, quantity, price…)
  · Withdrawal History  (columnas: Date(UTC), Coin, Network, Amount…)

NO soportado en fase 1 (futuros, earn, savings, convert):
  · Futures / Copy Trading → se detecta y se devuelve error descriptivo
  · Operaciones Earn / Staking / Savings
  · Conversiones (Convert)

Decisiones de diseño:
  · Importe siempre = Decimal(price) × Decimal(quantity) — NUNCA amount_usdt
    (amount_usdt es USD-equivalent, no la moneda quote real)
  · Cada fill de orden = una operación FIFO independiente (distinto timestamp)
  · Pares USDT: COMPRA/VENTA con contraparte USDT + advertencia suave
  · Fees MAKER = 0 exacto (correcto, no se trata como error)
  · Timezone naive: los timestamps de MEXC llevan el offset en el nombre de
    columna, no en el valor del string
"""

import logging
import warnings
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from datetime import datetime

import openpyxl

logger = logging.getLogger(__name__)


# ── EXCEPCIONES TIPIFICADAS ───────────────────────────────────────────────────

class MexcUnsupportedFormatError(ValueError):
    """
    El archivo MEXC es válido pero corresponde a un formato conscientemente
    no soportado (futuros, copy trading, staking, earn, etc.).

    No es un bug del parser. No debe computar como tasa de error en métricas.

    Atributos:
        code      — identificador del formato: "futures" | "staking" | "earn" |
                    "copy_trading" | "aggregated_history"
        category  — siempre "unsupported_format"
    """
    def __init__(self, code: str, mensaje_usuario: str):
        super().__init__(mensaje_usuario)
        self.code     = code
        self.category = "unsupported_format"


class MexcUserError(ValueError):
    """
    Error imputable al usuario: archivo incorrecto, XLSX corrupto, export vacío.

    No es un bug del parser. No debe provocar emails de soporte automáticos.

    Atributos:
        code      — identificador: "wrong_file" | "xlsx_corrupt" | "empty_file"
        category  — siempre "user_error"
    """
    def __init__(self, code: str, mensaje_usuario: str):
        super().__init__(mensaje_usuario)
        self.code     = code
        self.category = "user_error"


# ── FIRMAS DE DETECCIÓN ────────────────────────────────────────────────────────

MEXC_SPOT_SIGNATURE       = {"order_id", "symbol", "currency", "quantity", "price"}
MEXC_SPOT_ES_SIGNATURE    = {"pares", "dirección", "precio promedio completo", "cantidad completa"}
MEXC_WITHDRAWAL_SIGNATURE = {"coin", "network", "amount", "transactionfee", "status"}
MEXC_FUTURES_SIGNATURE    = {"futures", "vol(cont)", "deal_avg_price", "close_avg_price"}

# ── CONSTANTES FISCALES ───────────────────────────────────────────────────────

STABLES_EUR = {"EUR", "EURX"}
STABLES_USD = {"USDC", "USDT", "BUSD", "USD", "FDUSD", "DAI"}
STABLES     = STABLES_EUR | STABLES_USD

# ── FORMATOS DE FECHA ─────────────────────────────────────────────────────────

_FECHA_FORMATOS_SPOT = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
]

_FECHA_FORMATOS_WITHDRAWAL = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
]


# ── DATACLASSES (misma forma que clasificador.py para compatibilidad FIFO) ─────

@dataclass
class OperacionCompraventa:
    fecha: str
    tipo: str           # "COMPRA" | "VENTA"
    activo: str         # moneda base (ETH, BTC, TICS…)
    cantidad: float
    contraparte: str    # moneda quote (EUR, USDT…)
    importe: float
    fee_activo: str
    fee_cantidad: float

@dataclass
class OperacionMovimiento:
    fecha: str
    subtipo: str        # "Withdrawal"
    activo: str
    cantidad: float
    observacion: str

@dataclass
class OperacionRendimiento:
    fecha: str
    subtipo: str
    activo: str
    cantidad: float
    cuenta: str

@dataclass
class OperacionSwap:
    fecha: str
    activo_entregado: str
    cantidad_entregada: float
    activo_recibido: str
    cantidad_recibida: float
    nota: str = ""

@dataclass
class OperacionDesconocida:
    fecha: str
    subtipo: str
    activo: str
    cantidad: float
    cuenta: str


# ── HELPERS ───────────────────────────────────────────────────────────────────

def _leer_primera_hoja(filepath: str) -> tuple[list[str], list[dict]]:
    """
    Lee la primera hoja del XLSX y devuelve (headers, filas_como_dict).
    Los headers se devuelven tal como están en el archivo.
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)

    ws = wb[wb.sheetnames[0]]
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]

    if not rows:
        return [], []

    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        d = {}
        for col_i, h in enumerate(raw_headers):
            d[h] = str(row[col_i]).strip() if col_i < len(row) and row[col_i] is not None else ""
        data.append(d)

    return raw_headers, data


def _detectar_tipo_mexc(filepath: str) -> str:
    """
    Inspecciona los headers de la primera hoja y devuelve el tipo de export.
    Retorna: "spot" | "spot_es" | "withdrawal" | "futures" |
             "unknown" | "xlsx_corrupt" | "empty_file"
    """
    try:
        headers, data = _leer_primera_hoja(filepath)
    except Exception as e:
        logger.warning("MEXC: no se pudo leer el archivo: %s", e)
        return "xlsx_corrupt"

    if not headers:
        return "empty_file"

    headers_lower = {h.lower() for h in headers}

    if MEXC_FUTURES_SIGNATURE.issubset(headers_lower):
        return "futures"
    if MEXC_SPOT_SIGNATURE.issubset(headers_lower):
        return "spot"
    if MEXC_SPOT_ES_SIGNATURE.issubset(headers_lower):
        return "spot_es"
    if MEXC_WITHDRAWAL_SIGNATURE.issubset(headers_lower):
        return "withdrawal"

    return "unknown"


def _resolver_columna(row: dict, candidatos: list) -> str:
    """
    Busca el primer candidato que existe en row (case-insensitive).
    Devuelve el valor, o "" si no existe ninguno.
    """
    row_lower = {k.lower(): v for k, v in row.items()}
    for candidato in candidatos:
        if candidato.lower() in row_lower:
            return row_lower[candidato.lower()]
    return ""


def _parse_fecha(valor: str, formatos: list) -> str:
    """
    Parsea un string de fecha con los formatos dados.
    Devuelve "YYYY-MM-DD HH:MM:SS" o el string original si no coincide.
    """
    valor = valor.strip()
    for fmt in formatos:
        try:
            dt = datetime.strptime(valor, fmt)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    logger.warning("MEXC: fecha no reconocida: %r", valor)
    return valor


def _parse_decimal(valor: str) -> Decimal:
    """Convierte string a Decimal. Devuelve Decimal('0') si es inválido."""
    try:
        clean = valor.strip().replace(",", "")
        if not clean:
            return Decimal("0")
        return Decimal(clean)
    except InvalidOperation:
        return Decimal("0")


def _contar_filas_xlsx(filepath: str) -> int:
    """Cuenta las filas de datos del XLSX (sin contar la cabecera)."""
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
        ws = wb[wb.sheetnames[0]]
        rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
        return max(0, len(rows) - 1)
    except Exception:
        return 0


# ── CLASIFICADOR ──────────────────────────────────────────────────────────────

class ClasificadorMEXC:
    """
    Clasificador fiscal para exports XLS/XLSX de MEXC.

    Expone la misma interfaz que ClasificadorBinance / ClasificadorNexo para
    ser compatible con _pipeline_motor y procesar_con_fifo.

    Atributos públicos:
        compraventas  — BUY/SELL spot como OperacionCompraventa
        movimientos   — retiros como OperacionMovimiento
        rendimientos  — vacío en fase 1 (sin earn/staking)
        swaps         — vacío en fase 1 (sin convert)
        desconocidas  — operaciones no reconocidas
        advertencias  — warnings (pares USDT, operaciones ignoradas…)
        tipo_export   — "spot" | "withdrawal" (informativo)
    """

    # Columnas opcionales que pueden variar por versión de MEXC
    _FECHA_CANDIDATOS_SPOT = [
        "create_time(UTC+01:00)", "create_time(UTC+02:00)",
        "create_time(UTC+00:00)", "create_time(UTC)", "create_time", "time",
    ]
    _FECHA_CANDIDATOS_WITHDRAWAL = [
        "Date(UTC)", "date(utc)", "date",
    ]

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.tipo_export: str = ""

        self.compraventas:  list[OperacionCompraventa]  = []
        self.movimientos:   list[OperacionMovimiento]   = []
        self.rendimientos:  list[OperacionRendimiento]  = []
        self.swaps:         list[OperacionSwap]         = []
        self.desconocidas:  list[OperacionDesconocida]  = []
        self.advertencias:  list[str]                   = []

        # Flag: se activa si hay algún par USDT para el mensaje informativo
        self._tiene_pares_usdt: bool = False

    # ── ENTRADA PÚBLICA ────────────────────────────────────────────────────────

    def clasificar(self) -> "ClasificadorMEXC":
        tipo = _detectar_tipo_mexc(self.filepath)
        self.tipo_export = tipo

        if tipo == "xlsx_corrupt":
            raise MexcUserError(
                code="xlsx_corrupt",
                mensaje_usuario=(
                    "No hemos podido leer el archivo XLSX. "
                    "Intenta volver a descargarlo desde MEXC sin abrirlo con Excel."
                ),
            )
        if tipo == "empty_file":
            raise MexcUserError(
                code="empty_file",
                mensaje_usuario=(
                    "El archivo XLSX no contiene operaciones. "
                    "Exporta el rango de fechas con tus transacciones."
                ),
            )
        if tipo == "futures":
            raise MexcUnsupportedFormatError(
                code="futures",
                mensaje_usuario=(
                    "Este archivo corresponde a operaciones de futuros o copy trading de MEXC. "
                    "La app actualmente solo soporta spot. "
                    "Sube el archivo 'Trade Records' o 'Historial de Órdenes' (operaciones spot)."
                ),
            )
        if tipo == "spot":
            self._clasificar_spot()
        elif tipo == "spot_es":
            self._clasificar_spot_es()
        elif tipo == "withdrawal":
            self._clasificar_withdrawal()
        else:
            raise MexcUserError(
                code="wrong_file",
                mensaje_usuario=(
                    "El archivo no se reconoce como un export de MEXC. "
                    "Asegúrate de exportar el historial de Trade Records (operaciones spot), "
                    "el Historial de Órdenes o el historial de retiros desde tu cuenta MEXC."
                ),
            )

        if self._tiene_pares_usdt:
            self.advertencias.append(
                "Algunas operaciones están valoradas en USDT. "
                "Para la declaración del IRPF aplica el tipo de cambio EUR/USDT "
                "vigente en la fecha de cada operación."
            )

        if self.desconocidas:
            tipos = {d.subtipo for d in self.desconocidas}
            self.advertencias.append(
                f"{len(self.desconocidas)} operación(es) no reconocida(s): "
                + ", ".join(sorted(tipos))
            )

        return self

    # ── SPOT TRADES ───────────────────────────────────────────────────────────

    def _clasificar_spot(self) -> None:
        _, data = _leer_primera_hoja(self.filepath)
        if not data:
            return

        for fila in data:
            self._procesar_fila_spot(fila)

    def _procesar_fila_spot(self, fila: dict) -> None:
        fecha_raw = _resolver_columna(fila, self._FECHA_CANDIDATOS_SPOT)

        # Guardar filas con fecha vacía — pasarlas al motor causaría ValueError
        if not fecha_raw.strip():
            logger.warning("MEXC spot: fila con fecha vacía ignorada: %s", fila)
            self.desconocidas.append(OperacionDesconocida(
                fecha="", subtipo="fecha_vacia",
                activo=_resolver_columna(fila, ["symbol", "currency"]) or "?",
                cantidad=0.0, cuenta="MEXC",
            ))
            return

        fecha     = _parse_fecha(fecha_raw, _FECHA_FORMATOS_SPOT)

        symbol       = _resolver_columna(fila, ["symbol"])
        side         = _resolver_columna(fila, ["side"]).upper()
        quantity_raw = _resolver_columna(fila, ["quantity"])
        price_raw    = _resolver_columna(fila, ["price"])
        fee_raw      = _resolver_columna(fila, ["fee"])
        fee_cur      = _resolver_columna(fila, ["fee_currency"])
        trade_type   = _resolver_columna(fila, ["trade_type"]).upper()

        # Solo procesamos operaciones SPOT (ignorar futuros mezclados)
        if trade_type and trade_type not in ("SPOT", ""):
            self.desconocidas.append(OperacionDesconocida(
                fecha=fecha, subtipo=f"trade_type={trade_type}",
                activo=symbol, cantidad=0.0, cuenta="MEXC",
            ))
            return

        # Parsear símbolo BASE-QUOTE
        if "-" in symbol:
            base, _, quote = symbol.partition("-")
        else:
            base  = _resolver_columna(fila, ["currency"]) or symbol
            quote = "USDT"

        base  = base.strip().upper()
        quote = quote.strip().upper()

        qty    = _parse_decimal(quantity_raw)
        price  = _parse_decimal(price_raw)
        fee    = _parse_decimal(fee_raw)

        # Importe = price × quantity en la moneda quote (NUNCA amount_usdt)
        importe = qty * price

        if qty <= 0 or price <= 0:
            self.desconocidas.append(OperacionDesconocida(
                fecha=fecha, subtipo=f"qty/price inválidos ({side})",
                activo=base, cantidad=float(qty), cuenta="MEXC",
            ))
            return

        # Registrar par USDT para advertencia suave al final
        if quote in STABLES_USD and quote != "EUR":
            self._tiene_pares_usdt = True

        if side == "BUY":
            tipo = "COMPRA"
        elif side == "SELL":
            tipo = "VENTA"
        else:
            self.desconocidas.append(OperacionDesconocida(
                fecha=fecha, subtipo=f"side desconocido: {side!r}",
                activo=base, cantidad=float(qty), cuenta="MEXC",
            ))
            return

        self.compraventas.append(OperacionCompraventa(
            fecha      = fecha,
            tipo       = tipo,
            activo     = base,
            cantidad   = float(qty),
            contraparte= quote,
            importe    = float(importe),
            fee_activo = fee_cur.upper() if fee_cur else quote,
            fee_cantidad = float(fee),
        ))

    # ── SPOT ORDERS — formato español (Historial de Órdenes) ─────────────────
    #
    # Columnas: Pares · Tiempo · Tipo · Dirección · Precio Promedio Completo ·
    #           Precio de Orden · Cantidad Completa · Cantidad de Orden ·
    #           Monto de Orden · Estado
    #
    # "Cancelación parcial" = orden parcialmente ejecutada; los campos
    # Precio Promedio Completo y Cantidad Completa reflejan lo realmente
    # comprado/vendido → se procesa igual que Completado.

    _ESTADOS_SPOT_ES_VALIDOS = {"completado", "cancelación parcial", "cancelacion parcial"}

    def _clasificar_spot_es(self) -> None:
        _, data = _leer_primera_hoja(self.filepath)
        if not data:
            return
        for fila in data:
            self._procesar_fila_spot_es(fila)

    def _procesar_fila_spot_es(self, fila: dict) -> None:
        fecha_raw  = _resolver_columna(fila, ["Tiempo", "tiempo"])

        # Guardar filas con fecha vacía — pasarlas al motor causaría ValueError
        if not fecha_raw.strip():
            logger.warning("MEXC spot_es: fila con fecha vacía ignorada: %s", fila)
            self.desconocidas.append(OperacionDesconocida(
                fecha="", subtipo="fecha_vacia",
                activo=_resolver_columna(fila, ["Pares", "pares"]) or "?",
                cantidad=0.0, cuenta="MEXC",
            ))
            return

        fecha      = _parse_fecha(fecha_raw, _FECHA_FORMATOS_SPOT)

        pares       = _resolver_columna(fila, ["Pares", "pares"]).strip()
        direccion   = _resolver_columna(fila, ["Dirección", "dirección",
                                                "Direccion", "direccion"]).strip()
        precio_raw  = _resolver_columna(fila, ["Precio Promedio Completo",
                                                "precio promedio completo"])
        cant_raw    = _resolver_columna(fila, ["Cantidad Completa", "cantidad completa"])
        estado      = _resolver_columna(fila, ["Estado", "estado"]).strip()

        # Solo órdenes con fills reales
        if estado.lower() not in self._ESTADOS_SPOT_ES_VALIDOS:
            return

        # Parsear par BASE_QUOTE (separador "_", ej. "BNB_USDT")
        if "_" in pares:
            base, _, quote = pares.partition("_")
        else:
            base  = pares
            quote = "USDT"

        base  = base.strip().upper()
        quote = quote.strip().upper()

        qty   = _parse_decimal(cant_raw)
        price = _parse_decimal(precio_raw)

        if qty <= 0 or price <= 0:
            self.desconocidas.append(OperacionDesconocida(
                fecha=fecha, subtipo=f"qty/price inválidos ({direccion})",
                activo=base, cantidad=float(qty), cuenta="MEXC",
            ))
            return

        # Importe = price × quantity en moneda quote (misma regla que spot inglés)
        importe = qty * price

        # Registrar par USDT para advertencia suave
        if quote in STABLES_USD:
            self._tiene_pares_usdt = True

        dir_lower = direccion.lower()
        if dir_lower == "compra":
            tipo = "COMPRA"
        elif dir_lower == "venta":
            tipo = "VENTA"
        else:
            self.desconocidas.append(OperacionDesconocida(
                fecha=fecha, subtipo=f"dirección desconocida: {direccion!r}",
                activo=base, cantidad=float(qty), cuenta="MEXC",
            ))
            return

        # El formato español no tiene columna de fee → fee=0
        self.compraventas.append(OperacionCompraventa(
            fecha        = fecha,
            tipo         = tipo,
            activo       = base,
            cantidad     = float(qty),
            contraparte  = quote,
            importe      = float(importe),
            fee_activo   = quote,
            fee_cantidad = 0.0,
        ))

    # ── WITHDRAWAL HISTORY ────────────────────────────────────────────────────

    def _clasificar_withdrawal(self) -> None:
        _, data = _leer_primera_hoja(self.filepath)
        if not data:
            return

        for fila in data:
            self._procesar_fila_withdrawal(fila)

    def _procesar_fila_withdrawal(self, fila: dict) -> None:
        fecha_raw = _resolver_columna(fila, self._FECHA_CANDIDATOS_WITHDRAWAL)
        fecha     = _parse_fecha(fecha_raw, _FECHA_FORMATOS_WITHDRAWAL)

        coin       = _resolver_columna(fila, ["Coin", "coin"]).upper()
        network    = _resolver_columna(fila, ["Network", "network"])
        amount_raw = _resolver_columna(fila, ["Amount", "amount"])
        fee_raw    = _resolver_columna(fila, ["TransactionFee", "transactionfee", "fee"])
        txid       = _resolver_columna(fila, ["TXID", "txid"])
        status     = _resolver_columna(fila, ["Status", "status"]).strip()

        # Solo registrar retiros completados
        if status.lower() not in ("completed", ""):
            return

        amount = _parse_decimal(amount_raw)
        if amount <= 0:
            return

        fee    = _parse_decimal(fee_raw)
        obs    = f"Red: {network}" + (f" | TXID: {txid[:20]}…" if txid else "")
        if fee > 0:
            obs += f" | Fee: {float(fee):.8g} {coin}"

        self.movimientos.append(OperacionMovimiento(
            fecha      = fecha,
            subtipo    = "Withdrawal",
            activo     = coin,
            cantidad   = float(amount),
            observacion= obs,
        ))

    # ── RESUMEN ────────────────────────────────────────────────────────────────

    def resumen(self) -> dict:
        return {
            "tipo_export":  self.tipo_export,
            "compraventas": len(self.compraventas),
            "movimientos":  len(self.movimientos),
            "rendimientos": len(self.rendimientos),
            "swaps":        len(self.swaps),
            "desconocidas": len(self.desconocidas),
            "advertencias": len(self.advertencias),
        }


# ── EJECUCIÓN DIRECTA ─────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    ruta = sys.argv[1] if len(sys.argv) > 1 else "mexc.xlsx"
    print(f"\nProcesando: {ruta}\n")

    c = ClasificadorMEXC(ruta).clasificar()
    r = c.resumen()

    print("=" * 55)
    print(f"RESUMEN MEXC  —  tipo={c.tipo_export}")
    print("=" * 55)
    for k, v in r.items():
        print(f"  {k:<20} {v:>6}")

    print("\n── COMPRAVENTAS ──")
    for op in c.compraventas:
        print(f"  {op.fecha[:10]}  {op.tipo:<6}  {op.activo:<8}"
              f"  {op.cantidad:>14.8f}  @ {op.importe:.6f} {op.contraparte}"
              f"  fee: {op.fee_cantidad:.8g} {op.fee_activo}")

    print("\n── MOVIMIENTOS ──")
    for op in c.movimientos:
        print(f"  {op.fecha[:10]}  {op.subtipo:<15}  {op.activo:<6}"
              f"  {op.cantidad:>14.8f}  |  {op.observacion}")

    print(f"\n── ADVERTENCIAS ({len(c.advertencias)}) ──")
    for adv in c.advertencias:
        print(f"  ⚠  {adv}")
    if not c.advertencias:
        print("  Ninguna ✓")

    print("\n── DESCONOCIDAS ──")
    for op in c.desconocidas:
        print(f"  {op.fecha[:10]}  {op.subtipo:<30}  {op.activo:<8}  {op.cantidad:.6f}")
    if not c.desconocidas:
        print("  Ninguna ✓")
